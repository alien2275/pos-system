import os
import re
import secrets
import uuid
import hmac
import hashlib
import json
import zipfile
import shutil
from datetime import datetime
from decimal import Decimal, ROUND_HALF_UP
from io import BytesIO
from typing import Literal, Optional, List

from fastapi import FastAPI, File, Form, HTTPException, Request, Response, UploadFile
from fastapi.responses import JSONResponse
from openpyxl import load_workbook
from openpyxl import Workbook
from pydantic import BaseModel
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch
from reportlab.pdfgen import canvas
from sqlalchemy import create_engine, text
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles

app = FastAPI(title="POS API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://100.85.171.19:5173",
        "http://localhost:5173",
    ],
    allow_origin_regex=r"http://.*:5173",
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

UPLOAD_ROOT = os.getenv("UPLOAD_ROOT", "uploads")
PRODUCT_UPLOAD_DIR = os.path.join(UPLOAD_ROOT, "products")
EVENT_UPLOAD_DIR = os.path.join(UPLOAD_ROOT, "events")
ALLOWED_IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".webp", ".gif"}
MAX_IMAGE_BYTES = 5 * 1024 * 1024

os.makedirs(PRODUCT_UPLOAD_DIR, exist_ok=True)
os.makedirs(EVENT_UPLOAD_DIR, exist_ok=True)
app.mount("/uploads", StaticFiles(directory=UPLOAD_ROOT), name="uploads")

DB_USER = os.getenv("POSTGRES_USER")
DB_PASSWORD = os.getenv("POSTGRES_PASSWORD")
DB_NAME = os.getenv("POSTGRES_DB")
DB_HOST = os.getenv("POSTGRES_HOST", "postgres")

DATABASE_URL = f"postgresql://{DB_USER}:{DB_PASSWORD}@{DB_HOST}:5432/{DB_NAME}"

engine = create_engine(DATABASE_URL)

ADMIN_USERNAME = os.getenv("ADMIN_USERNAME", "admin")
ADMIN_PASSWORD = os.getenv("ADMIN_PASSWORD", "admin")
ADMIN_SESSION_SECRET = os.getenv(
    "ADMIN_SESSION_SECRET",
    DB_PASSWORD or secrets.token_urlsafe(32),
)
ADMIN_COOKIE_NAME = "pos_admin_session"


DEFAULT_SETTINGS = {
    "tax_enabled": "true",
    "tax_state": "MD",
    "tax_rate_percent": "6.00",
    "flat_shipping_cents": "600",
    "store_url": "http://100.85.171.19:5173/store",
    "pos_rounding_mode": "none",
}


def create_admin_session_token():
    signature = hmac.new(
        ADMIN_SESSION_SECRET.encode("utf-8"),
        b"admin",
        hashlib.sha256,
    ).hexdigest()
    return f"admin.{signature}"


def is_valid_admin_session(token):
    if not token:
        return False

    return hmac.compare_digest(token, create_admin_session_token())


def is_public_request(method, path):
    if method == "OPTIONS":
        return True

    if path in {"/", "/auth/login", "/auth/status", "/settings/store-qr.png"}:
        return True

    if method == "GET" and path == "/settings":
        return True

    if method == "GET" and re.match(r"^/events/\d+/images$", path):
        return True

    return (
        path.startswith("/store")
        or path.startswith("/uploads")
        or path.startswith("/docs")
        or path.startswith("/openapi.json")
    )


@app.middleware("http")
async def require_admin_session(request: Request, call_next):
    if is_public_request(request.method, request.url.path):
        return await call_next(request)

    if is_valid_admin_session(request.cookies.get(ADMIN_COOKIE_NAME)):
        return await call_next(request)

    return JSONResponse({"detail": "Authentication required"}, status_code=401)


def generate_order_number(conn):
    alphabet = "ABCDEFGHJKLMNPQRSTUVWXYZ23456789"

    for _ in range(10):
        order_number = "".join(secrets.choice(alphabet) for _ in range(8))
        existing = conn.execute(
            text("SELECT id FROM sales WHERE order_number = :order_number;"),
            {"order_number": order_number},
        ).first()

        if existing is None:
            return order_number

    raise HTTPException(status_code=500, detail="Could not generate order number")


def get_app_settings(conn):
    rows = conn.execute(text("SELECT key, value FROM app_settings;"))
    settings = DEFAULT_SETTINGS.copy()
    settings.update({row._mapping["key"]: row._mapping["value"] for row in rows})

    tax_enabled = str(settings["tax_enabled"]).lower() in {"1", "true", "yes", "on"}
    tax_rate = Decimal(settings["tax_rate_percent"])
    flat_shipping_cents = int(settings["flat_shipping_cents"])

    return {
        "tax_enabled": tax_enabled,
        "tax_state": settings["tax_state"],
        "tax_rate_percent": str(tax_rate.quantize(Decimal("0.01"))),
        "flat_shipping_cents": flat_shipping_cents,
        "store_url": settings["store_url"],
        "pos_rounding_mode": settings["pos_rounding_mode"],
    }


def calculate_tax_cents(amount_cents, tax_rate_percent):
    rate = Decimal(str(tax_rate_percent))
    tax = (Decimal(amount_cents) * rate / Decimal("100")).quantize(
        Decimal("1"),
        rounding=ROUND_HALF_UP,
    )
    return int(tax)


def calculate_rounding_adjustment_cents(total_cents, rounding_mode):
    if rounding_mode == "nearest_0_05":
        increment = 5
    elif rounding_mode == "nearest_0_10":
        increment = 10
    elif rounding_mode == "dollar_threshold_0_10":
        cents = total_cents % 100
        if cents == 0:
            return 0
        if cents <= 10:
            return -cents
        return 100 - cents
    else:
        return 0

    remainder = total_cents % increment
    if remainder == 0:
        return 0
    if remainder < increment / 2:
        return -remainder
    return increment - remainder


def format_money(cents):
    return f"${(int(cents or 0) / 100):,.2f}"


def draw_pdf_row(pdf, y, label, value, bold=False):
    pdf.setFont("Helvetica-Bold" if bold else "Helvetica", 10)
    pdf.drawString(0.75 * inch, y, label)
    pdf.drawRightString(7.75 * inch, y, str(value))
    return y - 0.24 * inch


def workbook_response(workbook, filename):
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def normalize_import_header(value):
    return str(value or "").strip().lower()


def clean_cell(value):
    if value is None:
        return None

    if isinstance(value, str):
        value = value.strip()
        return value or None

    return value


def normalize_sku(value):
    value = clean_cell(value)
    return value.upper() if isinstance(value, str) else value


def parse_import_money(value, row_number, field_name):
    value = clean_cell(value)

    if value is None:
        return 0

    try:
        if isinstance(value, str):
            value = value.replace("$", "").replace(",", "")

        return int(round(float(value) * 100))
    except (TypeError, ValueError):
        raise ValueError(f"Row {row_number}: {field_name} must be a number")


def parse_import_int(value, row_number, field_name):
    value = clean_cell(value)

    if value is None:
        return 0

    try:
        return int(value)
    except (TypeError, ValueError):
        raise ValueError(f"Row {row_number}: {field_name} must be a whole number")


def find_product_duplicate(conn, sku, barcode):
    if not sku and not barcode:
        return None

    return conn.execute(
        text(
            """
            SELECT *
            FROM products
            WHERE (:sku IS NOT NULL AND sku = :sku)
               OR (:barcode IS NOT NULL AND barcode = :barcode)
            ORDER BY is_active DESC, id ASC
            LIMIT 1;
            """
        ),
        {"sku": sku, "barcode": barcode},
    ).first()


def generate_import_sku(conn, base_sku, product_name):
    clean_base = str(base_sku or product_name or "ITEM").strip().upper()

    trailing_number = re.match(r"^(.*?)(\d+)$", clean_base)

    if trailing_number:
        prefix = trailing_number.group(1)
        starting_number = int(trailing_number.group(2))
        padding = len(trailing_number.group(2))

        for number in range(starting_number + 1, 10000):
            candidate = f"{prefix}{number:0{padding}d}"
            existing = conn.execute(
                text("SELECT id FROM products WHERE LOWER(sku) = LOWER(:sku) LIMIT 1;"),
                {"sku": candidate},
            ).first()

            if existing is None:
                return candidate

    clean_base = "".join(
        character if character.isalnum() else "-"
        for character in clean_base
    ).strip("-")
    clean_base = clean_base or "ITEM"

    for number in range(1, 1000):
        candidate = f"{clean_base}-{number:03d}"
        existing = conn.execute(
            text("SELECT id FROM products WHERE LOWER(sku) = LOWER(:sku) LIMIT 1;"),
            {"sku": candidate},
        ).first()

        if existing is None:
            return candidate

    raise ValueError(f"Could not generate a new SKU for {clean_base}")


def insert_import_product(conn, product_data):
    product_data = {
        **product_data,
        "sku": normalize_sku(product_data.get("sku")),
    }

    conn.execute(
        text(
            """
            INSERT INTO products
            (
                sku,
                barcode,
                name,
                category,
                description,
                price_cents,
                cost_cents,
                quantity_on_hand,
                reorder_level,
                image_url,
                public_description,
                is_public
            )
            VALUES
            (
                :sku,
                :barcode,
                :name,
                :category,
                :description,
                :price_cents,
                :cost_cents,
                :quantity_on_hand,
                :reorder_level,
                :image_url,
                :public_description,
                :is_public
            );
            """
        ),
        product_data,
    )


def update_import_product(conn, product_id, product_data):
    product_data = {
        **product_data,
        "sku": normalize_sku(product_data.get("sku")),
    }

    conn.execute(
        text(
            """
            UPDATE products
            SET sku = :sku,
                barcode = :barcode,
                name = :name,
                category = :category,
                description = :description,
                price_cents = :price_cents,
                cost_cents = :cost_cents,
                quantity_on_hand = :quantity_on_hand,
                reorder_level = :reorder_level,
                public_description = :public_description,
                updated_at = CURRENT_TIMESTAMP
            WHERE id = :id;
            """
        ),
        {
            **product_data,
            "id": product_id,
        },
    )


class ProductCreate(BaseModel):
    sku: Optional[str] = None
    barcode: Optional[str] = None
    name: str
    category: Optional[str] = None
    description: Optional[str] = None
    price_cents: int = 0
    cost_cents: int = 0
    quantity_on_hand: int = 0
    reorder_level: int = 0
    image_url: Optional[str] = None
    public_description: Optional[str] = None
    is_public: bool = False

    def model_post_init(self, __context):
        self.sku = normalize_sku(self.sku)


class ProductUpdate(BaseModel):
    sku: Optional[str] = None
    barcode: Optional[str] = None
    name: Optional[str] = None
    category: Optional[str] = None
    description: Optional[str] = None
    price_cents: Optional[int] = None
    cost_cents: Optional[int] = None
    quantity_on_hand: Optional[int] = None
    reorder_level: Optional[int] = None
    is_active: Optional[bool] = None
    image_url: Optional[str] = None
    public_description: Optional[str] = None
    is_public: Optional[bool] = None

    def model_post_init(self, __context):
        self.sku = normalize_sku(self.sku)


class ImportDuplicateDecision(BaseModel):
    action: Literal["skip", "update", "import_as_new"]
    product_data: dict
    matched_product_id: Optional[int] = None


class ImportDuplicateResolution(BaseModel):
    decisions: List[ImportDuplicateDecision]


class InventoryAdjustment(BaseModel):
    product_id: int
    quantity_change: int
    reason: Literal[
        "Shipment",
        "Sale",
        "Return",
        "Damage",
        "Adjustment",
        "Transfer"
    ]
    notes: Optional[str] = None


class SaleItemCreate(BaseModel):
    product_id: int
    quantity: int


class SaleCreate(BaseModel):
    items: List[SaleItemCreate]
    customer_name: Optional[str] = None
    payment_type: Optional[Literal["cash", "card", "other"]] = None


class AdminLogin(BaseModel):
    username: str
    password: str


class AppSettingsUpdate(BaseModel):
    tax_enabled: Optional[bool] = None
    tax_state: Optional[str] = None
    tax_rate_percent: Optional[Decimal] = None
    flat_shipping_cents: Optional[int] = None
    store_url: Optional[str] = None
    pos_rounding_mode: Optional[
        Literal["none", "nearest_0_05", "nearest_0_10", "dollar_threshold_0_10"]
    ] = None


class EventCreate(BaseModel):
    title: str
    location: Optional[str] = None
    description: Optional[str] = None
    start_date: str
    end_date: Optional[str] = None
    image_url: Optional[str] = None
    is_public: bool = True


class EventUpdate(BaseModel):
    title: Optional[str] = None
    location: Optional[str] = None
    description: Optional[str] = None
    start_date: Optional[str] = None
    end_date: Optional[str] = None
    image_url: Optional[str] = None
    is_public: Optional[bool] = None


class OnlineOrderItemCreate(BaseModel):
    product_id: int
    quantity: int


class OnlineOrderCreate(BaseModel):
    customer_name: str
    customer_email: str
    shipping_name: str
    shipping_address_line1: str
    shipping_address_line2: Optional[str] = None
    shipping_city: str
    shipping_state: str
    shipping_postal_code: str
    shipping_country: str = "US"
    payment_provider: str = "placeholder"
    payment_reference: Optional[str] = None
    items: List[OnlineOrderItemCreate]


class OnlineOrderShipmentUpdate(BaseModel):
    carrier: Literal["USPS", "UPS", "FedEx", "Other"]
    tracking_id: str


@app.on_event("startup")
def ensure_event_table():
    with engine.begin() as conn:
        conn.execute(
            text(
                """
                CREATE TABLE IF NOT EXISTS app_settings (
                    key TEXT PRIMARY KEY,
                    value TEXT NOT NULL,
                    updated_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP
                );
                """
            )
        )
        for key, value in DEFAULT_SETTINGS.items():
            conn.execute(
                text(
                    """
                    INSERT INTO app_settings (key, value)
                    VALUES (:key, :value)
                    ON CONFLICT (key) DO NOTHING;
                    """
                ),
                {"key": key, "value": value},
            )
        conn.execute(
            text("ALTER TABLE sales ADD COLUMN IF NOT EXISTS order_number TEXT;")
        )
        conn.execute(text("ALTER TABLE sales ADD COLUMN IF NOT EXISTS customer_name TEXT;"))
        conn.execute(text("ALTER TABLE sales ADD COLUMN IF NOT EXISTS payment_type TEXT;"))
        conn.execute(text("ALTER TABLE sales ADD COLUMN IF NOT EXISTS subtotal_cents INTEGER;"))
        conn.execute(text("ALTER TABLE sales ADD COLUMN IF NOT EXISTS tax_cents INTEGER NOT NULL DEFAULT 0;"))
        conn.execute(text("ALTER TABLE sales ADD COLUMN IF NOT EXISTS tax_rate_percent NUMERIC(8, 3) NOT NULL DEFAULT 0;"))
        conn.execute(text("ALTER TABLE sales ADD COLUMN IF NOT EXISTS rounding_adjustment_cents INTEGER NOT NULL DEFAULT 0;"))
        conn.execute(
            text(
                """
                UPDATE sales
                SET subtotal_cents = total_cents
                WHERE subtotal_cents IS NULL;
                """
            )
        )
        conn.execute(
            text(
                """
                CREATE TABLE IF NOT EXISTS events (
                    id SERIAL PRIMARY KEY,
                    title TEXT NOT NULL,
                    location TEXT,
                    description TEXT,
                    start_date DATE NOT NULL,
                    end_date DATE,
                    image_url TEXT,
                    is_public BOOLEAN NOT NULL DEFAULT TRUE,
                    created_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,
                    updated_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP
                );
                """
            )
        )
        conn.execute(
            text(
                """
                CREATE TABLE IF NOT EXISTS event_images (
                    id SERIAL PRIMARY KEY,
                    event_id INTEGER NOT NULL REFERENCES events(id) ON DELETE CASCADE,
                    image_url TEXT NOT NULL,
                    created_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP
                );
                """
            )
        )
        conn.execute(
            text(
                """
                CREATE TABLE IF NOT EXISTS product_images (
                    id SERIAL PRIMARY KEY,
                    product_id INTEGER NOT NULL REFERENCES products(id) ON DELETE CASCADE,
                    image_url TEXT NOT NULL,
                    created_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP
                );
                """
            )
        )
        conn.execute(
            text(
                """
                CREATE TABLE IF NOT EXISTS online_orders (
                    id SERIAL PRIMARY KEY,
                    customer_name TEXT,
                    customer_email TEXT,
                    shipping_name TEXT,
                    shipping_address_line1 TEXT,
                    shipping_address_line2 TEXT,
                    shipping_city TEXT,
                    shipping_state TEXT,
                    shipping_postal_code TEXT,
                    shipping_country TEXT,
                    payment_provider TEXT,
                    payment_reference TEXT,
                    sale_id INTEGER REFERENCES sales(id),
                    carrier TEXT,
                    tracking_id TEXT,
                    status TEXT NOT NULL DEFAULT 'pending_packaging',
                    subtotal_cents INTEGER NOT NULL DEFAULT 0,
                    tax_cents INTEGER NOT NULL DEFAULT 0,
                    shipping_cents INTEGER NOT NULL DEFAULT 0,
                    tax_rate_percent NUMERIC(8, 3) NOT NULL DEFAULT 0,
                    total_cents INTEGER NOT NULL DEFAULT 0,
                    packaged_at TIMESTAMP,
                    shipped_at TIMESTAMP,
                    archived_at TIMESTAMP,
                    created_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,
                    updated_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP
                );
                """
            )
        )
        conn.execute(text("ALTER TABLE online_orders ADD COLUMN IF NOT EXISTS carrier TEXT;"))
        conn.execute(text("ALTER TABLE online_orders ADD COLUMN IF NOT EXISTS tracking_id TEXT;"))
        conn.execute(text("ALTER TABLE online_orders ADD COLUMN IF NOT EXISTS sale_id INTEGER REFERENCES sales(id);"))
        conn.execute(text("ALTER TABLE online_orders ADD COLUMN IF NOT EXISTS packaged_at TIMESTAMP;"))
        conn.execute(text("ALTER TABLE online_orders ADD COLUMN IF NOT EXISTS shipped_at TIMESTAMP;"))
        conn.execute(text("ALTER TABLE online_orders ADD COLUMN IF NOT EXISTS archived_at TIMESTAMP;"))
        conn.execute(text("ALTER TABLE online_orders ADD COLUMN IF NOT EXISTS subtotal_cents INTEGER NOT NULL DEFAULT 0;"))
        conn.execute(text("ALTER TABLE online_orders ADD COLUMN IF NOT EXISTS tax_cents INTEGER NOT NULL DEFAULT 0;"))
        conn.execute(text("ALTER TABLE online_orders ADD COLUMN IF NOT EXISTS shipping_cents INTEGER NOT NULL DEFAULT 0;"))
        conn.execute(text("ALTER TABLE online_orders ADD COLUMN IF NOT EXISTS tax_rate_percent NUMERIC(8, 3) NOT NULL DEFAULT 0;"))
        conn.execute(
            text(
                """
                UPDATE online_orders
                SET subtotal_cents = total_cents
                WHERE subtotal_cents = 0
                  AND total_cents > 0
                  AND tax_cents = 0
                  AND shipping_cents = 0;
                """
            )
        )
        conn.execute(
            text(
                """
                UPDATE sales
                SET order_number = SUBSTRING(UPPER(MD5('sale-' || id::text || '-' || created_at::text)) FROM 1 FOR 8)
                WHERE order_number IS NULL
                   OR order_number LIKE 'POS-%'
                   OR order_number LIKE 'WEB-%';
                """
            )
        )
        conn.execute(
            text(
                "ALTER TABLE online_orders ALTER COLUMN status SET DEFAULT 'pending_packaging';"
            )
        )
        conn.execute(
            text(
                """
                UPDATE online_orders
                SET status = 'pending_packaging'
                WHERE status = 'pending_fulfillment';
                """
            )
        )
        conn.execute(
            text(
                """
                CREATE TABLE IF NOT EXISTS online_order_items (
                    id SERIAL PRIMARY KEY,
                    order_id INTEGER NOT NULL REFERENCES online_orders(id) ON DELETE CASCADE,
                    product_id INTEGER REFERENCES products(id),
                    product_name TEXT NOT NULL,
                    quantity INTEGER NOT NULL,
                    price_cents INTEGER NOT NULL
                );
                """
            )
        )
        existing_online_orders = conn.execute(
            text(
                """
                SELECT *
                FROM online_orders
                WHERE sale_id IS NULL;
                """
            )
        )

        for order in existing_online_orders:
            order_data = dict(order._mapping)
            order_number = generate_order_number(conn)
            sale = conn.execute(
                text(
                    """
                    INSERT INTO sales (
                        subtotal_cents,
                        tax_cents,
                        tax_rate_percent,
                        total_cents,
                        order_number,
                        created_at
                    )
                    VALUES (
                        :subtotal_cents,
                        :tax_cents,
                        :tax_rate_percent,
                        :total_cents,
                        :order_number,
                        :created_at
                    )
                    RETURNING *;
                    """
                ),
                {
                    "subtotal_cents": order_data.get("subtotal_cents") or order_data["total_cents"],
                    "tax_cents": order_data.get("tax_cents") or 0,
                    "tax_rate_percent": order_data.get("tax_rate_percent") or 0,
                    "total_cents": order_data["total_cents"],
                    "order_number": order_number,
                    "created_at": order_data["created_at"],
                },
            ).first()
            sale_id = sale._mapping["id"]

            conn.execute(
                text(
                    """
                    UPDATE online_orders
                    SET sale_id = :sale_id,
                        updated_at = CURRENT_TIMESTAMP
                    WHERE id = :id;
                    """
                ),
                {"id": order_data["id"], "sale_id": sale_id},
            )

            order_items = conn.execute(
                text(
                    """
                    SELECT *
                    FROM online_order_items
                    WHERE order_id = :order_id;
                    """
                ),
                {"order_id": order_data["id"]},
            )

            for item in order_items:
                item_data = dict(item._mapping)
                conn.execute(
                    text(
                        """
                        INSERT INTO sale_items
                        (sale_id, product_id, quantity, price_cents)
                        VALUES
                        (:sale_id, :product_id, :quantity, :price_cents);
                        """
                    ),
                    {
                        "sale_id": sale_id,
                        "product_id": item_data["product_id"],
                        "quantity": item_data["quantity"],
                        "price_cents": item_data["price_cents"],
                    },
                )
    

@app.get("/")
def root():
    return {
        "status": "online",
        "service": "POS API"
    }


@app.get("/auth/status")
def auth_status(request: Request):
    return {
        "authenticated": is_valid_admin_session(
            request.cookies.get(ADMIN_COOKIE_NAME)
        )
    }


@app.post("/auth/login")
def login(login_data: AdminLogin):
    username_matches = hmac.compare_digest(login_data.username, ADMIN_USERNAME)
    password_matches = hmac.compare_digest(login_data.password, ADMIN_PASSWORD)

    if not username_matches or not password_matches:
        raise HTTPException(status_code=401, detail="Invalid username or password")

    response = JSONResponse({"authenticated": True})
    response.set_cookie(
        key=ADMIN_COOKIE_NAME,
        value=create_admin_session_token(),
        httponly=True,
        samesite="lax",
        max_age=60 * 60 * 12,
    )
    return response


@app.post("/auth/logout")
def logout():
    response = JSONResponse({"authenticated": False})
    response.delete_cookie(ADMIN_COOKIE_NAME)
    return response


@app.get("/settings")
def get_settings():
    with engine.connect() as conn:
        return get_app_settings(conn)


@app.put("/settings")
def update_settings(settings_update: AppSettingsUpdate):
    fields = settings_update.model_dump(exclude_unset=True)

    if not fields:
        raise HTTPException(status_code=400, detail="No settings provided")

    if "tax_rate_percent" in fields:
        tax_rate = Decimal(fields["tax_rate_percent"])
        if tax_rate < 0 or tax_rate > 25:
            raise HTTPException(status_code=400, detail="Tax rate must be between 0 and 25")
        fields["tax_rate_percent"] = str(tax_rate.quantize(Decimal("0.01")))

    if "tax_enabled" in fields:
        fields["tax_enabled"] = "true" if fields["tax_enabled"] else "false"

    if "flat_shipping_cents" in fields and fields["flat_shipping_cents"] < 0:
        raise HTTPException(status_code=400, detail="Flat shipping cannot be negative")

    if "tax_state" in fields:
        fields["tax_state"] = fields["tax_state"].strip().upper()[:20] or "MD"

    if "store_url" in fields:
        store_url = fields["store_url"].strip()
        if not store_url.startswith(("http://", "https://")):
            raise HTTPException(
                status_code=400,
                detail="Store URL must start with http:// or https://",
            )
        fields["store_url"] = store_url

    if "pos_rounding_mode" in fields:
        allowed_rounding_modes = {
            "none",
            "nearest_0_05",
            "nearest_0_10",
            "dollar_threshold_0_10",
        }
        if fields["pos_rounding_mode"] not in allowed_rounding_modes:
            raise HTTPException(status_code=400, detail="Invalid POS rounding mode")

    with engine.begin() as conn:
        for key, value in fields.items():
            conn.execute(
                text(
                    """
                    INSERT INTO app_settings (key, value, updated_at)
                    VALUES (:key, :value, CURRENT_TIMESTAMP)
                    ON CONFLICT (key) DO UPDATE
                    SET value = EXCLUDED.value,
                        updated_at = CURRENT_TIMESTAMP;
                    """
                ),
                {"key": key, "value": str(value)},
            )

        return get_app_settings(conn)


@app.get("/settings/store-qr.png")
def get_store_qr_code(url: Optional[str] = None):
    if url:
        store_url = url.strip()
        if not store_url.startswith(("http://", "https://")):
            raise HTTPException(
                status_code=400,
                detail="Store URL must start with http:// or https://",
            )
    else:
        with engine.connect() as conn:
            store_url = get_app_settings(conn)["store_url"]

    import qrcode

    qr_image = qrcode.make(store_url)
    buffer = BytesIO()
    qr_image.save(buffer, format="PNG")
    buffer.seek(0)

    return Response(
        buffer.getvalue(),
        media_type="image/png",
        headers={"Cache-Control": "no-store"},
    )


@app.get("/reports/tax-summary.pdf")
def download_tax_summary_report(start_date: str, end_date: str):
    with engine.connect() as conn:
        summary = conn.execute(
            text(
                """
                SELECT
                    COUNT(*) AS transaction_count,
                    COALESCE(SUM(COALESCE(sales.subtotal_cents, sales.total_cents)), 0) AS taxable_sales_cents,
                    COALESCE(SUM(sales.tax_cents), 0) AS tax_collected_cents,
                    COALESCE(SUM(COALESCE(online_orders.shipping_cents, 0)), 0) AS shipping_collected_cents,
                    COALESCE(SUM(sales.total_cents), 0) AS gross_receipts_cents
                FROM sales
                LEFT JOIN online_orders ON online_orders.sale_id = sales.id
                WHERE sales.created_at >= CAST(:start_date AS date)
                  AND sales.created_at < (CAST(:end_date AS date) + INTERVAL '1 day');
                """
            ),
            {"start_date": start_date, "end_date": end_date},
        ).first()

        by_type = conn.execute(
            text(
                """
                SELECT
                    CASE WHEN online_orders.id IS NULL THEN 'POS' ELSE 'Online' END AS sale_type,
                    COUNT(*) AS transaction_count,
                    COALESCE(SUM(COALESCE(sales.subtotal_cents, sales.total_cents)), 0) AS taxable_sales_cents,
                    COALESCE(SUM(sales.tax_cents), 0) AS tax_collected_cents,
                    COALESCE(SUM(COALESCE(online_orders.shipping_cents, 0)), 0) AS shipping_collected_cents,
                    COALESCE(SUM(sales.total_cents), 0) AS gross_receipts_cents
                FROM sales
                LEFT JOIN online_orders ON online_orders.sale_id = sales.id
                WHERE sales.created_at >= CAST(:start_date AS date)
                  AND sales.created_at < (CAST(:end_date AS date) + INTERVAL '1 day')
                GROUP BY sale_type
                ORDER BY sale_type;
                """
            ),
            {"start_date": start_date, "end_date": end_date},
        )

        by_rate = conn.execute(
            text(
                """
                SELECT
                    COALESCE(sales.tax_rate_percent, 0) AS tax_rate_percent,
                    COUNT(*) AS transaction_count,
                    COALESCE(SUM(COALESCE(sales.subtotal_cents, sales.total_cents)), 0) AS taxable_sales_cents,
                    COALESCE(SUM(sales.tax_cents), 0) AS tax_collected_cents
                FROM sales
                WHERE sales.created_at >= CAST(:start_date AS date)
                  AND sales.created_at < (CAST(:end_date AS date) + INTERVAL '1 day')
                GROUP BY tax_rate_percent
                ORDER BY tax_rate_percent;
                """
            ),
            {"start_date": start_date, "end_date": end_date},
        )

    buffer = BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter
    y = height - 0.75 * inch

    pdf.setFont("Helvetica-Bold", 18)
    pdf.drawString(0.75 * inch, y, "sammyinthesky Tax Summary")
    y -= 0.32 * inch
    pdf.setFont("Helvetica", 10)
    pdf.drawString(0.75 * inch, y, f"Reporting period: {start_date} through {end_date}")
    y -= 0.2 * inch
    pdf.drawString(0.75 * inch, y, "Generated from POS and online store records.")
    y -= 0.36 * inch

    pdf.setStrokeColor(colors.lightgrey)
    pdf.line(0.75 * inch, y, width - 0.75 * inch, y)
    y -= 0.35 * inch

    summary_data = dict(summary._mapping)
    pdf.setFont("Helvetica-Bold", 13)
    pdf.drawString(0.75 * inch, y, "Summary")
    y -= 0.3 * inch
    y = draw_pdf_row(pdf, y, "Transactions", summary_data["transaction_count"])
    y = draw_pdf_row(pdf, y, "Taxable sales", format_money(summary_data["taxable_sales_cents"]))
    y = draw_pdf_row(pdf, y, "Sales tax collected", format_money(summary_data["tax_collected_cents"]), True)
    y = draw_pdf_row(pdf, y, "Shipping collected", format_money(summary_data["shipping_collected_cents"]))
    y = draw_pdf_row(pdf, y, "Gross receipts", format_money(summary_data["gross_receipts_cents"]), True)
    y -= 0.18 * inch

    pdf.setFont("Helvetica-Bold", 13)
    pdf.drawString(0.75 * inch, y, "Breakdown By Sale Type")
    y -= 0.3 * inch
    for row in by_type:
        row_data = dict(row._mapping)
        y = draw_pdf_row(pdf, y, f"{row_data['sale_type']} transactions", row_data["transaction_count"])
        y = draw_pdf_row(pdf, y, f"{row_data['sale_type']} taxable sales", format_money(row_data["taxable_sales_cents"]))
        y = draw_pdf_row(pdf, y, f"{row_data['sale_type']} tax collected", format_money(row_data["tax_collected_cents"]))
        y = draw_pdf_row(pdf, y, f"{row_data['sale_type']} shipping collected", format_money(row_data["shipping_collected_cents"]))
        y -= 0.08 * inch

    y -= 0.12 * inch
    pdf.setFont("Helvetica-Bold", 13)
    pdf.drawString(0.75 * inch, y, "Breakdown By Tax Rate")
    y -= 0.3 * inch
    for row in by_rate:
        row_data = dict(row._mapping)
        rate = Decimal(row_data["tax_rate_percent"] or 0).quantize(Decimal("0.01"))
        y = draw_pdf_row(
            pdf,
            y,
            f"{rate}% - {row_data['transaction_count']} transactions",
            f"{format_money(row_data['taxable_sales_cents'])} taxable / {format_money(row_data['tax_collected_cents'])} tax",
        )

    y -= 0.25 * inch
    pdf.setFont("Helvetica", 8)
    pdf.setFillColor(colors.darkgrey)
    pdf.drawString(
        0.75 * inch,
        y,
        "For recordkeeping only. Confirm filing requirements with Maryland Tax Connect or a tax professional.",
    )

    pdf.showPage()
    pdf.save()
    buffer.seek(0)

    filename = f"tax-summary-{start_date}-to-{end_date}.pdf"
    return Response(
        buffer.getvalue(),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/exports/products.xlsx")
def export_products():
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Products"
    worksheet.append(
        [
            "SKU",
            "Barcode",
            "Name",
            "Category",
            "Description",
            "Public Store Description",
            "Price",
            "Cost",
            "Quantity",
            "Reorder Level",
            "Public",
            "Image URL",
            "Active",
        ]
    )

    with engine.connect() as conn:
        rows = conn.execute(
            text(
                """
                SELECT *
                FROM products
                ORDER BY is_active DESC, category NULLS LAST, name ASC;
                """
            )
        )

        for row in rows:
            product = dict(row._mapping)
            worksheet.append(
                [
                    product["sku"],
                    product["barcode"],
                    product["name"],
                    product["category"],
                    product["description"],
                    product["public_description"],
                    (product["price_cents"] or 0) / 100,
                    (product["cost_cents"] or 0) / 100,
                    product["quantity_on_hand"],
                    product["reorder_level"],
                    product["is_public"],
                    product["image_url"],
                    product["is_active"],
                ]
            )

    return workbook_response(workbook, "products-export.xlsx")


@app.get("/exports/sales.xlsx")
def export_sales():
    workbook = Workbook()
    sales_sheet = workbook.active
    sales_sheet.title = "Sales"
    sales_sheet.append(
        [
            "Order Number",
            "Type",
            "Customer",
            "Payment Type",
            "Subtotal",
            "Tax",
            "Rounding",
            "Total",
            "Created At",
        ]
    )

    items_sheet = workbook.create_sheet("Sale Items")
    items_sheet.append(
        [
            "Order Number",
            "Product",
            "SKU",
            "Quantity",
            "Price Each",
            "Line Total",
        ]
    )

    with engine.connect() as conn:
        sales = conn.execute(
            text(
                """
                SELECT
                    sales.*,
                    CASE WHEN online_orders.id IS NULL THEN 'POS' ELSE 'Online' END AS sale_type,
                    COALESCE(
                        sales.customer_name,
                        online_orders.customer_name,
                        online_orders.shipping_name
                    ) AS display_customer_name
                FROM sales
                LEFT JOIN online_orders ON online_orders.sale_id = sales.id
                ORDER BY sales.created_at DESC;
                """
            )
        )

        for row in sales:
            sale = dict(row._mapping)
            sales_sheet.append(
                [
                    sale["order_number"],
                    sale["sale_type"],
                    sale["display_customer_name"],
                    sale["payment_type"],
                    (sale["subtotal_cents"] or sale["total_cents"] or 0) / 100,
                    (sale["tax_cents"] or 0) / 100,
                    (sale["rounding_adjustment_cents"] or 0) / 100,
                    (sale["total_cents"] or 0) / 100,
                    sale["created_at"],
                ]
            )

        sale_items = conn.execute(
            text(
                """
                SELECT
                    sales.order_number,
                    products.name,
                    products.sku,
                    sale_items.quantity,
                    sale_items.price_cents
                FROM sale_items
                JOIN sales ON sales.id = sale_items.sale_id
                LEFT JOIN products ON products.id = sale_items.product_id
                ORDER BY sales.created_at DESC, sale_items.id ASC;
                """
            )
        )

        for row in sale_items:
            item = dict(row._mapping)
            items_sheet.append(
                [
                    item["order_number"],
                    item["name"],
                    item["sku"],
                    item["quantity"],
                    (item["price_cents"] or 0) / 100,
                    ((item["price_cents"] or 0) * item["quantity"]) / 100,
                ]
            )

    return workbook_response(workbook, "sales-export.xlsx")


@app.get("/exports/orders.xlsx")
def export_online_orders():
    workbook = Workbook()
    orders_sheet = workbook.active
    orders_sheet.title = "Online Orders"
    orders_sheet.append(
        [
            "Order Number",
            "Status",
            "Customer",
            "Email",
            "Ship To",
            "Address 1",
            "Address 2",
            "City",
            "State",
            "Postal Code",
            "Country",
            "Subtotal",
            "Tax",
            "Shipping",
            "Total",
            "Carrier",
            "Tracking",
            "Created At",
            "Shipped At",
            "Archived At",
        ]
    )

    items_sheet = workbook.create_sheet("Order Items")
    items_sheet.append(
        ["Order Number", "Product", "Quantity", "Price Each", "Line Total"]
    )

    with engine.connect() as conn:
        orders = conn.execute(
            text(
                """
                SELECT
                    online_orders.*,
                    sales.order_number
                FROM online_orders
                LEFT JOIN sales ON sales.id = online_orders.sale_id
                ORDER BY online_orders.created_at DESC;
                """
            )
        )

        for row in orders:
            order = dict(row._mapping)
            orders_sheet.append(
                [
                    order["order_number"],
                    order["status"],
                    order["customer_name"],
                    order["customer_email"],
                    order["shipping_name"],
                    order["shipping_address_line1"],
                    order["shipping_address_line2"],
                    order["shipping_city"],
                    order["shipping_state"],
                    order["shipping_postal_code"],
                    order["shipping_country"],
                    (order["subtotal_cents"] or 0) / 100,
                    (order["tax_cents"] or 0) / 100,
                    (order["shipping_cents"] or 0) / 100,
                    (order["total_cents"] or 0) / 100,
                    order["carrier"],
                    order["tracking_id"],
                    order["created_at"],
                    order["shipped_at"],
                    order["archived_at"],
                ]
            )

        items = conn.execute(
            text(
                """
                SELECT
                    sales.order_number,
                    online_order_items.product_name,
                    online_order_items.quantity,
                    online_order_items.price_cents
                FROM online_order_items
                JOIN online_orders ON online_orders.id = online_order_items.order_id
                LEFT JOIN sales ON sales.id = online_orders.sale_id
                ORDER BY online_orders.created_at DESC, online_order_items.id ASC;
                """
            )
        )

        for row in items:
            item = dict(row._mapping)
            items_sheet.append(
                [
                    item["order_number"],
                    item["product_name"],
                    item["quantity"],
                    (item["price_cents"] or 0) / 100,
                    ((item["price_cents"] or 0) * item["quantity"]) / 100,
                ]
            )

    return workbook_response(workbook, "online-orders-export.xlsx")


@app.get("/backups/full.zip")
def download_full_backup():
    backup_tables = [
        "app_settings",
        "products",
        "product_images",
        "inventory_transactions",
        "sales",
        "sale_items",
        "events",
        "event_images",
        "online_orders",
        "online_order_items",
    ]
    created_at = datetime.utcnow().replace(microsecond=0).isoformat() + "Z"
    backup_data = {
        "created_at": created_at,
        "database": {},
    }

    with engine.connect() as conn:
        for table_name in backup_tables:
            order_column = "key" if table_name == "app_settings" else "id"
            rows = conn.execute(
                text(f"SELECT * FROM {table_name} ORDER BY {order_column};")
            )
            backup_data["database"][table_name] = [
                dict(row._mapping) for row in rows
            ]

    buffer = BytesIO()
    with zipfile.ZipFile(buffer, "w", compression=zipfile.ZIP_DEFLATED) as backup_zip:
        backup_zip.writestr(
            "database.json",
            json.dumps(backup_data, default=str, indent=2),
        )
        backup_zip.writestr(
            "README.txt",
            (
                "POS System Backup\n\n"
                f"Created at: {created_at}\n\n"
                "database.json contains exported table data.\n"
                "uploads/ contains product and event image files.\n"
                "Restore should be handled carefully to avoid overwriting newer sales.\n"
            ),
        )

        if os.path.isdir(UPLOAD_ROOT):
            for root, _, files in os.walk(UPLOAD_ROOT):
                for filename in files:
                    disk_path = os.path.join(root, filename)
                    archive_path = os.path.join(
                        "uploads",
                        os.path.relpath(disk_path, UPLOAD_ROOT),
                    )
                    backup_zip.write(disk_path, archive_path)

    buffer.seek(0)
    filename = f"pos-backup-{datetime.utcnow().strftime('%Y%m%d-%H%M%S')}.zip"

    return Response(
        content=buffer.getvalue(),
        media_type="application/zip",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.post("/backups/restore")
async def restore_full_backup(
    confirm: str = Form(...),
    file: UploadFile = File(...),
):
    if confirm != "RESTORE BACKUP":
        raise HTTPException(
            status_code=400,
            detail='Type "RESTORE BACKUP" to restore from a backup ZIP',
        )

    contents = await file.read()

    try:
        backup_zip = zipfile.ZipFile(BytesIO(contents))
    except zipfile.BadZipFile:
        raise HTTPException(status_code=400, detail="Backup file must be a valid ZIP")

    if "database.json" not in backup_zip.namelist():
        raise HTTPException(status_code=400, detail="Backup ZIP is missing database.json")

    try:
        backup_data = json.loads(backup_zip.read("database.json").decode("utf-8"))
    except (json.JSONDecodeError, UnicodeDecodeError):
        raise HTTPException(status_code=400, detail="database.json is invalid")

    database = backup_data.get("database")
    if not isinstance(database, dict):
        raise HTTPException(status_code=400, detail="Backup database payload is invalid")

    restore_tables = [
        "app_settings",
        "products",
        "product_images",
        "inventory_transactions",
        "sales",
        "sale_items",
        "events",
        "event_images",
        "online_orders",
        "online_order_items",
    ]
    delete_order = [
        "online_order_items",
        "online_orders",
        "event_images",
        "events",
        "sale_items",
        "inventory_transactions",
        "product_images",
        "products",
        "sales",
        "app_settings",
    ]
    insert_order = [
        "app_settings",
        "products",
        "sales",
        "product_images",
        "inventory_transactions",
        "sale_items",
        "events",
        "event_images",
        "online_orders",
        "online_order_items",
    ]

    with engine.begin() as conn:
        for table_name in delete_order:
            conn.execute(text(f"DELETE FROM {table_name};"))

        for table_name in insert_order:
            rows = database.get(table_name, [])
            if not rows:
                continue

            for row in rows:
                columns = list(row.keys())
                column_list = ", ".join(columns)
                value_list = ", ".join(f":{column}" for column in columns)
                conn.execute(
                    text(
                        f"""
                        INSERT INTO {table_name} ({column_list})
                        VALUES ({value_list});
                        """
                    ),
                    row,
                )

        for table_name in restore_tables:
            if table_name == "app_settings":
                continue

            conn.execute(
                text(
                    """
                    SELECT setval(
                        pg_get_serial_sequence(:table_name, 'id'),
                        COALESCE((SELECT MAX(id) FROM """ + table_name + """), 1),
                        true
                    );
                    """
                ),
                {"table_name": table_name},
            )

    if os.path.isdir(UPLOAD_ROOT):
        shutil.rmtree(UPLOAD_ROOT)
    os.makedirs(PRODUCT_UPLOAD_DIR, exist_ok=True)
    os.makedirs(EVENT_UPLOAD_DIR, exist_ok=True)

    for name in backup_zip.namelist():
        if not name.startswith("uploads/") or name.endswith("/"):
            continue

        relative_path = name[len("uploads/"):]
        target_path = os.path.abspath(os.path.join(UPLOAD_ROOT, relative_path))
        upload_root_abs = os.path.abspath(UPLOAD_ROOT)

        if os.path.commonpath([upload_root_abs, target_path]) != upload_root_abs:
            raise HTTPException(status_code=400, detail="Backup contains unsafe path")

        os.makedirs(os.path.dirname(target_path), exist_ok=True)
        with open(target_path, "wb") as output_file:
            output_file.write(backup_zip.read(name))

    return {
        "restored": True,
        "tables": list(database.keys()),
        "created_at": backup_data.get("created_at"),
    }


@app.get("/products")
def get_products():
    with engine.connect() as conn:
        result = conn.execute(
            text(
                """
                SELECT *
                FROM products
                WHERE is_active = TRUE
                ORDER BY id;
                """
            )
        )

        return [
            dict(row._mapping)
            for row in result
        ]


@app.get("/store/products")
def get_store_products():
    with engine.connect() as conn:
        result = conn.execute(
            text(
                """
                SELECT
                    id,
                    sku,
                    name,
                    category,
                    public_description,
                    image_url,
                    price_cents,
                    quantity_on_hand
                FROM products
                WHERE is_active = TRUE
                  AND is_public = TRUE
                ORDER BY category NULLS LAST, name ASC;
                """
            )
        )

        products = [dict(row._mapping) for row in result]

        if not products:
            return []

        images = conn.execute(
            text(
                """
                SELECT *
                FROM product_images
                WHERE product_id IN (
                    SELECT id
                    FROM products
                    WHERE is_active = TRUE
                      AND is_public = TRUE
                )
                ORDER BY created_at ASC, id ASC;
                """
            )
        )

        images_by_product = {}
        for row in images:
            image = dict(row._mapping)
            images_by_product.setdefault(image["product_id"], []).append(image)

        return [
            {
                **product,
                "images": images_by_product.get(product["id"], []),
            }
            for product in products
        ]


@app.post("/store/orders")
def create_store_order(order: OnlineOrderCreate):
    if not order.items:
        raise HTTPException(status_code=400, detail="Order must contain at least one item")

    with engine.begin() as conn:
        settings = get_app_settings(conn)
        subtotal_cents = 0
        order_items = []

        for item in order.items:
            if item.quantity <= 0:
                raise HTTPException(status_code=400, detail="Quantity must be greater than zero")

            product = conn.execute(
                text(
                    """
                    SELECT *
                    FROM products
                    WHERE id = :id
                      AND is_active = TRUE
                      AND is_public = TRUE;
                    """
                ),
                {"id": item.product_id},
            ).first()

            if product is None:
                raise HTTPException(status_code=404, detail="Product not found")

            product_data = dict(product._mapping)

            if product_data["quantity_on_hand"] < item.quantity:
                raise HTTPException(
                    status_code=400,
                    detail=f"Not enough inventory for {product_data['name']}",
                )

            line_total = product_data["price_cents"] * item.quantity
            subtotal_cents += line_total
            order_items.append(
                {
                    "product_id": item.product_id,
                    "product_name": product_data["name"],
                    "quantity": item.quantity,
                    "price_cents": product_data["price_cents"],
                }
            )

        shipping_cents = settings["flat_shipping_cents"]
        tax_cents = (
            calculate_tax_cents(subtotal_cents, settings["tax_rate_percent"])
            if settings["tax_enabled"]
            else 0
        )
        tax_rate_percent = settings["tax_rate_percent"] if settings["tax_enabled"] else 0
        total_cents = subtotal_cents + tax_cents + shipping_cents

        order_row = conn.execute(
            text(
                """
                INSERT INTO online_orders
                (
                    customer_name,
                    customer_email,
                    shipping_name,
                    shipping_address_line1,
                    shipping_address_line2,
                    shipping_city,
                    shipping_state,
                    shipping_postal_code,
                    shipping_country,
                    payment_provider,
                    payment_reference,
                    status,
                    subtotal_cents,
                    tax_cents,
                    shipping_cents,
                    tax_rate_percent,
                    total_cents
                )
                VALUES
                (
                    :customer_name,
                    :customer_email,
                    :shipping_name,
                    :shipping_address_line1,
                    :shipping_address_line2,
                    :shipping_city,
                    :shipping_state,
                    :shipping_postal_code,
                    :shipping_country,
                    :payment_provider,
                    :payment_reference,
                    'pending_packaging',
                    :subtotal_cents,
                    :tax_cents,
                    :shipping_cents,
                    :tax_rate_percent,
                    :total_cents
                )
                RETURNING *;
                """
            ),
            {
                **order.model_dump(exclude={"items"}),
                "subtotal_cents": subtotal_cents,
                "tax_cents": tax_cents,
                "shipping_cents": shipping_cents,
                "tax_rate_percent": tax_rate_percent,
                "total_cents": total_cents,
            },
        ).first()

        order_id = order_row._mapping["id"]
        order_number = generate_order_number(conn)
        sale_row = conn.execute(
            text(
                """
                INSERT INTO sales (
                    customer_name,
                    payment_type,
                    subtotal_cents,
                    tax_cents,
                    tax_rate_percent,
                    total_cents,
                    order_number
                )
                VALUES (
                    :customer_name,
                    :payment_type,
                    :subtotal_cents,
                    :tax_cents,
                    :tax_rate_percent,
                    :total_cents,
                    :order_number
                )
                RETURNING *;
                """
            ),
            {
                "customer_name": order.customer_name,
                "payment_type": order.payment_provider,
                "subtotal_cents": subtotal_cents,
                "tax_cents": tax_cents,
                "tax_rate_percent": tax_rate_percent,
                "total_cents": total_cents,
                "order_number": order_number,
            },
        ).first()
        sale_id = sale_row._mapping["id"]

        conn.execute(
            text(
                """
                UPDATE online_orders
                SET sale_id = :sale_id,
                    updated_at = CURRENT_TIMESTAMP
                WHERE id = :id;
                """
            ),
            {"id": order_id, "sale_id": sale_id},
        )

        for item in order_items:
            conn.execute(
                text(
                    """
                    INSERT INTO online_order_items
                    (order_id, product_id, product_name, quantity, price_cents)
                    VALUES
                    (:order_id, :product_id, :product_name, :quantity, :price_cents);
                    """
                ),
                {"order_id": order_id, **item},
            )

            conn.execute(
                text(
                    """
                    INSERT INTO sale_items
                    (sale_id, product_id, quantity, price_cents)
                    VALUES
                    (:sale_id, :product_id, :quantity, :price_cents);
                    """
                ),
                {
                    "sale_id": sale_id,
                    "product_id": item["product_id"],
                    "quantity": item["quantity"],
                    "price_cents": item["price_cents"],
                },
            )

            conn.execute(
                text(
                    """
                    UPDATE products
                    SET quantity_on_hand = quantity_on_hand - :quantity,
                        updated_at = CURRENT_TIMESTAMP
                    WHERE id = :product_id;
                    """
                ),
                item,
            )

            conn.execute(
                text(
                    """
                    INSERT INTO inventory_transactions
                    (product_id, quantity_change, reason, notes)
                    VALUES
                    (:product_id, :quantity_change, 'Sale', :notes);
                    """
                ),
                {
                    "product_id": item["product_id"],
                    "quantity_change": -item["quantity"],
                    "notes": f"Online order #{order_id}",
                },
            )

    return {
        "order": {
            **dict(order_row._mapping),
            "sale_id": sale_id,
            "order_number": sale_row._mapping["order_number"],
        },
        "items": order_items,
    }


@app.get("/store/events")
def get_store_events():
    with engine.connect() as conn:
        result = conn.execute(
            text(
                """
                SELECT *
                FROM events
                WHERE is_public = TRUE
                  AND COALESCE(end_date, start_date) >= CURRENT_DATE
                ORDER BY start_date ASC, title ASC;
                """
            )
        )

        return [dict(row._mapping) for row in result]


@app.get("/store/past-events")
def get_store_past_events():
    with engine.connect() as conn:
        result = conn.execute(
            text(
                """
                SELECT *
                FROM events
                WHERE is_public = TRUE
                  AND COALESCE(end_date, start_date) < CURRENT_DATE
                ORDER BY start_date DESC, title ASC;
                """
            )
        )

        return [dict(row._mapping) for row in result]


@app.get("/events")
def get_events():
    with engine.connect() as conn:
        result = conn.execute(
            text(
                """
                SELECT *
                FROM events
                ORDER BY start_date DESC, title ASC;
                """
            )
        )

        return [dict(row._mapping) for row in result]


@app.get("/online-orders")
def get_online_orders(include_archived: bool = False):
    with engine.connect() as conn:
        orders = conn.execute(
            text(
                """
                SELECT
                    online_orders.*,
                    sales.order_number
                FROM online_orders
                LEFT JOIN sales ON sales.id = online_orders.sale_id
                WHERE (:include_archived = TRUE OR archived_at IS NULL)
                ORDER BY
                    CASE status
                        WHEN 'pending_packaging' THEN 1
                        WHEN 'packaged' THEN 2
                        WHEN 'shipped' THEN 3
                        ELSE 4
                    END,
                    created_at DESC;
                """
            ),
            {"include_archived": include_archived},
        )

        order_rows = [dict(row._mapping) for row in orders]

        if not order_rows:
            return []

        items = conn.execute(
            text(
                """
                SELECT *
                FROM online_order_items
                WHERE order_id IN (
                    SELECT id
                    FROM online_orders
                )
                ORDER BY id;
                """
            )
        )

        items_by_order = {}
        for row in items:
            item = dict(row._mapping)
            items_by_order.setdefault(item["order_id"], []).append(item)

        return [
            {
                **order,
                "items": items_by_order.get(order["id"], []),
            }
            for order in order_rows
        ]


@app.get("/online-orders/{order_id}")
def get_online_order(order_id: int):
    with engine.connect() as conn:
        order = conn.execute(
            text(
                """
                SELECT
                    online_orders.*,
                    sales.order_number
                FROM online_orders
                LEFT JOIN sales ON sales.id = online_orders.sale_id
                WHERE online_orders.id = :id;
                """
            ),
            {"id": order_id},
        ).first()

        if order is None:
            raise HTTPException(status_code=404, detail="Online order not found")

        items = conn.execute(
            text(
                """
                SELECT *
                FROM online_order_items
                WHERE order_id = :order_id
                ORDER BY id;
                """
            ),
            {"order_id": order_id},
        )

        return {
            **dict(order._mapping),
            "items": [dict(row._mapping) for row in items],
        }


@app.put("/online-orders/{order_id}/packaged")
def mark_online_order_packaged(order_id: int):
    with engine.begin() as conn:
        result = conn.execute(
            text(
                """
                UPDATE online_orders
                SET status = 'packaged',
                    packaged_at = COALESCE(packaged_at, CURRENT_TIMESTAMP),
                    updated_at = CURRENT_TIMESTAMP
                WHERE id = :id
                RETURNING *;
                """
            ),
            {"id": order_id},
        ).first()

    if result is None:
        raise HTTPException(status_code=404, detail="Online order not found")

    return dict(result._mapping)


@app.put("/online-orders/{order_id}/ship")
def mark_online_order_shipped(order_id: int, shipment: OnlineOrderShipmentUpdate):
    fields = shipment.model_dump()

    if not fields["tracking_id"].strip():
        raise HTTPException(status_code=400, detail="Tracking ID is required")

    with engine.begin() as conn:
        result = conn.execute(
            text(
                """
                UPDATE online_orders
                SET status = 'shipped',
                    carrier = :carrier,
                    tracking_id = :tracking_id,
                    packaged_at = COALESCE(packaged_at, CURRENT_TIMESTAMP),
                    shipped_at = CURRENT_TIMESTAMP,
                    updated_at = CURRENT_TIMESTAMP
                WHERE id = :id
                RETURNING *;
                """
            ),
            {"id": order_id, **fields},
        ).first()

    if result is None:
        raise HTTPException(status_code=404, detail="Online order not found")

    return dict(result._mapping)


@app.put("/online-orders/{order_id}/archive")
def archive_online_order(order_id: int):
    with engine.begin() as conn:
        result = conn.execute(
            text(
                """
                UPDATE online_orders
                SET archived_at = CURRENT_TIMESTAMP,
                    updated_at = CURRENT_TIMESTAMP
                WHERE id = :id
                  AND status = 'shipped'
                RETURNING *;
                """
            ),
            {"id": order_id},
        ).first()

    if result is None:
        raise HTTPException(
            status_code=404,
            detail="Shipped online order not found",
        )

    return dict(result._mapping)


@app.get("/events/{event_id}/images")
def get_event_images(event_id: int):
    with engine.connect() as conn:
        event = conn.execute(
            text("SELECT * FROM events WHERE id = :id;"),
            {"id": event_id},
        ).first()

        if event is None:
            raise HTTPException(status_code=404, detail="Event not found")

        result = conn.execute(
            text(
                """
                SELECT *
                FROM event_images
                WHERE event_id = :event_id
                ORDER BY created_at DESC, id DESC;
                """
            ),
            {"event_id": event_id},
        )

        return [dict(row._mapping) for row in result]


@app.post("/events")
def create_event(event: EventCreate):
    with engine.begin() as conn:
        result = conn.execute(
            text(
                """
                INSERT INTO events
                (
                    title,
                    location,
                    description,
                    start_date,
                    end_date,
                    image_url,
                    is_public
                )
                VALUES
                (
                    :title,
                    :location,
                    :description,
                    :start_date,
                    :end_date,
                    :image_url,
                    :is_public
                )
                RETURNING *;
                """
            ),
            event.model_dump(),
        ).first()

    return dict(result._mapping)


@app.put("/events/{event_id}")
def update_event(event_id: int, event: EventUpdate):
    fields = event.model_dump(exclude_unset=True)

    if not fields:
        raise HTTPException(status_code=400, detail="No update fields provided")

    set_clause = ", ".join([f"{key} = :{key}" for key in fields.keys()])

    query = f"""
        UPDATE events
        SET {set_clause},
            updated_at = CURRENT_TIMESTAMP
        WHERE id = :id
        RETURNING *;
    """

    fields["id"] = event_id

    with engine.begin() as conn:
        result = conn.execute(text(query), fields).first()

    if result is None:
        raise HTTPException(status_code=404, detail="Event not found")

    return dict(result._mapping)


@app.post("/events/{event_id}/image")
async def upload_event_image(event_id: int, file: UploadFile = File(...)):
    with engine.connect() as conn:
        event = conn.execute(
            text("SELECT * FROM events WHERE id = :id;"),
            {"id": event_id},
        ).first()

    if event is None:
        raise HTTPException(status_code=404, detail="Event not found")

    _, extension = os.path.splitext(file.filename or "")
    extension = extension.lower()

    if extension not in ALLOWED_IMAGE_EXTENSIONS:
        raise HTTPException(
            status_code=400,
            detail="Image must be a JPG, PNG, WebP, or GIF file",
        )

    if not file.content_type or not file.content_type.startswith("image/"):
        raise HTTPException(status_code=400, detail="Uploaded file must be an image")

    contents = await file.read()

    if len(contents) > MAX_IMAGE_BYTES:
        raise HTTPException(status_code=400, detail="Image must be 5 MB or smaller")

    filename = f"event-{event_id}-{uuid.uuid4().hex}{extension}"
    disk_path = os.path.join(EVENT_UPLOAD_DIR, filename)
    image_url = f"/uploads/events/{filename}"

    with open(disk_path, "wb") as image_file:
        image_file.write(contents)

    with engine.begin() as conn:
        updated_event = conn.execute(
            text(
                """
                UPDATE events
                SET image_url = :image_url,
                    updated_at = CURRENT_TIMESTAMP
                WHERE id = :id
                RETURNING *;
                """
            ),
            {"id": event_id, "image_url": image_url},
        ).first()

    return dict(updated_event._mapping)


@app.post("/events/{event_id}/images")
async def upload_event_gallery_image(event_id: int, file: UploadFile = File(...)):
    with engine.connect() as conn:
        event = conn.execute(
            text("SELECT * FROM events WHERE id = :id;"),
            {"id": event_id},
        ).first()

    if event is None:
        raise HTTPException(status_code=404, detail="Event not found")

    _, extension = os.path.splitext(file.filename or "")
    extension = extension.lower()

    if extension not in ALLOWED_IMAGE_EXTENSIONS:
        raise HTTPException(
            status_code=400,
            detail="Image must be a JPG, PNG, WebP, or GIF file",
        )

    if not file.content_type or not file.content_type.startswith("image/"):
        raise HTTPException(status_code=400, detail="Uploaded file must be an image")

    contents = await file.read()

    if len(contents) > MAX_IMAGE_BYTES:
        raise HTTPException(status_code=400, detail="Image must be 5 MB or smaller")

    filename = f"event-{event_id}-gallery-{uuid.uuid4().hex}{extension}"
    disk_path = os.path.join(EVENT_UPLOAD_DIR, filename)
    image_url = f"/uploads/events/{filename}"

    with open(disk_path, "wb") as image_file:
        image_file.write(contents)

    with engine.begin() as conn:
        event_image = conn.execute(
            text(
                """
                INSERT INTO event_images (event_id, image_url)
                VALUES (:event_id, :image_url)
                RETURNING *;
                """
            ),
            {"event_id": event_id, "image_url": image_url},
        ).first()

    return dict(event_image._mapping)


@app.delete("/event-images/{image_id}")
def delete_event_image(image_id: int):
    with engine.begin() as conn:
        result = conn.execute(
            text(
                """
                DELETE FROM event_images
                WHERE id = :id
                RETURNING *;
                """
            ),
            {"id": image_id},
        ).first()

    if result is None:
        raise HTTPException(status_code=404, detail="Event image not found")

    return {"deleted": True, "image": dict(result._mapping)}


@app.delete("/events/{event_id}")
def delete_event(event_id: int):
    with engine.begin() as conn:
        result = conn.execute(
            text(
                """
                DELETE FROM events
                WHERE id = :id
                RETURNING *;
                """
            ),
            {"id": event_id},
        ).first()

    if result is None:
        raise HTTPException(status_code=404, detail="Event not found")

    return {"deleted": True, "event": dict(result._mapping)}


@app.get("/products/search/{query}")
def search_products(query: str):
    search_term = f"%{query}%"

    with engine.connect() as conn:
        result = conn.execute(
            text(
                """
                SELECT *
                FROM products
                WHERE
                    name ILIKE :search
                    OR sku ILIKE :search
                    OR barcode ILIKE :search
                    OR description ILIKE :search
                ORDER BY name;
                """
            ),
            {"search": search_term},
        )

        return [
            dict(row._mapping)
            for row in result
        ]
    

@app.get("/products/inactive-match")
def get_inactive_product_match(sku: str = "", barcode: str = ""):
    with engine.connect() as conn:
        result = conn.execute(
            text(
                """
                SELECT *
                FROM products
                WHERE is_active = FALSE
                  AND (
                    (:sku <> '' AND sku = :sku)
                    OR (:barcode <> '' AND barcode = :barcode)
                  )
                LIMIT 1;
                """
            ),
            {"sku": sku, "barcode": barcode},
        ).first()

    if result is None:
        raise HTTPException(status_code=404, detail="No inactive product found")

    return dict(result._mapping)


@app.get("/products/import-template")
def get_product_import_template():
    headers = [
        "sku",
        "barcode",
        "name",
        "category",
        "description",
        "public store description",
        "price",
        "cost",
        "quantity",
        "reorder level",
    ]

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Products"
    sheet.append(headers)
    sheet.append(
        [
            "EARRING-001",
            "123456789012",
            "Example Earrings",
            "Jewelry",
            "Internal product notes",
            "Public storefront description",
            "12.50",
            "4.25",
            "3",
            "1",
        ]
    )

    output = BytesIO()
    workbook.save(output)

    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": 'attachment; filename="product-import-template.xlsx"'
        },
    )


@app.post("/products/import")
async def import_products(file: UploadFile = File(...)):
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Upload an .xlsx file")

    contents = await file.read()

    try:
        workbook = load_workbook(filename=BytesIO(contents), data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail="Could not read Excel file") from exc

    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))

    if not rows:
        raise HTTPException(status_code=400, detail="Excel file is empty")

    headers = [normalize_import_header(value) for value in rows[0]]
    required_headers = [
        "sku",
        "barcode",
        "name",
        "category",
        "description",
        "public store description",
        "price",
        "cost",
        "quantity",
        "reorder level",
    ]

    missing_headers = [header for header in required_headers if header not in headers]

    if missing_headers:
        raise HTTPException(
            status_code=400,
            detail=f"Missing columns: {', '.join(missing_headers)}",
        )

    header_indexes = {
        header: headers.index(header)
        for header in required_headers
    }

    imported = 0
    updated = 0
    skipped = 0
    errors = []
    generated_skus = []
    duplicates = []

    with engine.begin() as conn:
        for index, row in enumerate(rows[1:], start=2):
            if not any(clean_cell(value) is not None for value in row):
                continue

            try:
                name = clean_cell(row[header_indexes["name"]])

                if not name:
                    raise ValueError(f"Row {index}: name is required")

                product_data = {
                    "sku": normalize_sku(row[header_indexes["sku"]]),
                    "barcode": clean_cell(row[header_indexes["barcode"]]),
                    "name": name,
                    "category": clean_cell(row[header_indexes["category"]]),
                    "description": clean_cell(row[header_indexes["description"]]),
                    "public_description": clean_cell(
                        row[header_indexes["public store description"]]
                    ),
                    "price_cents": parse_import_money(
                        row[header_indexes["price"]],
                        index,
                        "price",
                    ),
                    "cost_cents": parse_import_money(
                        row[header_indexes["cost"]],
                        index,
                        "cost",
                    ),
                    "quantity_on_hand": parse_import_int(
                        row[header_indexes["quantity"]],
                        index,
                        "quantity",
                    ),
                    "reorder_level": parse_import_int(
                        row[header_indexes["reorder level"]],
                        index,
                        "reorder level",
                    ),
                    "image_url": None,
                    "is_public": False,
                }

                duplicate = find_product_duplicate(
                    conn,
                    product_data["sku"],
                    product_data["barcode"],
                )

                if duplicate is not None:
                    duplicate_data = dict(duplicate._mapping)
                    duplicates.append(
                        {
                            "row_number": index,
                            "product_data": product_data,
                            "matched_product": {
                                "id": duplicate_data["id"],
                                "sku": duplicate_data["sku"],
                                "barcode": duplicate_data["barcode"],
                                "name": duplicate_data["name"],
                                "category": duplicate_data["category"],
                                "quantity_on_hand": duplicate_data["quantity_on_hand"],
                            },
                        }
                    )
                    continue

                insert_import_product(conn, product_data)
                imported += 1
            except Exception as exc:
                errors.append(str(exc))

    return {
        "imported": imported,
        "updated": updated,
        "skipped": skipped,
        "generated_skus": generated_skus,
        "duplicates": duplicates,
        "errors": errors,
    }


@app.post("/products/import/resolve-duplicates")
def resolve_product_import_duplicates(resolution: ImportDuplicateResolution):
    imported = 0
    updated = 0
    skipped = 0
    errors = []
    generated_skus = []

    with engine.begin() as conn:
        for index, decision in enumerate(resolution.decisions, start=1):
            try:
                product_data = decision.product_data

                if decision.action == "skip":
                    skipped += 1
                    continue

                if decision.action == "update":
                    if not decision.matched_product_id:
                        raise ValueError(f"Decision {index}: matched product is required")

                    update_import_product(
                        conn,
                        decision.matched_product_id,
                        product_data,
                    )
                    updated += 1
                    continue

                if decision.action == "import_as_new":
                    generated_sku = generate_import_sku(
                        conn,
                        product_data.get("sku"),
                        product_data.get("name"),
                    )
                    generated_skus.append(
                        f"{product_data.get('sku') or product_data.get('name')} imported as {generated_sku}"
                    )
                    product_data = {
                        **product_data,
                        "sku": generated_sku,
                        "barcode": None,
                    }
                    insert_import_product(conn, product_data)
                    imported += 1
            except Exception as exc:
                errors.append(str(exc))

    return {
        "imported": imported,
        "updated": updated,
        "skipped": skipped,
        "generated_skus": generated_skus,
        "errors": errors,
    }


@app.get("/products/barcode/{barcode}")
def get_product_by_barcode(barcode: str):
    with engine.connect() as conn:
        result = conn.execute(
            text(
                """
                SELECT *
                FROM products
                WHERE barcode = :barcode
                """
            ),
            {"barcode": barcode},
        ).first()

    if result is None:
        raise HTTPException(
            status_code=404,
            detail="Product not found"
        )

    return dict(result._mapping)


@app.get("/products/sku/{sku}")
def get_product_by_sku(sku: str):
    with engine.connect() as conn:
        result = conn.execute(
            text(
                """
                SELECT *
                FROM products
                WHERE sku = :sku
                """
            ),
            {"sku": sku},
        ).first()

    if result is None:
        raise HTTPException(
            status_code=404,
            detail="Product not found"
        )

    return dict(result._mapping)


@app.get("/products/{product_id}")
def get_product_by_id(product_id: int):
    with engine.connect() as conn:
        result = conn.execute(
            text(
                """
                SELECT *
                FROM products
                WHERE id = :id
                """
            ),
            {"id": product_id},
        ).first()

    if result is None:
        raise HTTPException(
            status_code=404,
            detail="Product not found"
        )

    return dict(result._mapping)


@app.get("/categories")
def get_categories():
    with engine.connect() as conn:
        result = conn.execute(
            text(
                """
                SELECT DISTINCT category
                FROM products
                WHERE category IS NOT NULL
                  AND category <> ''
                ORDER BY category;
                """
            )
        )

        return [row[0] for row in result]
    

@app.get("/products/category/{category}")
def get_products_by_category(category: str):
    with engine.connect() as conn:
        result = conn.execute(
            text(
                """
                SELECT *
                FROM products
                WHERE category = :category
                ORDER BY name;
                """
            ),
            {"category": category},
        )

        return [
            dict(row._mapping)
            for row in result
        ]
    

@app.get("/dashboard")
def get_dashboard():
    with engine.connect() as conn:
        product_summary = conn.execute(
            text(
                """
                SELECT
                    COUNT(*) AS product_count,
                    COUNT(*) FILTER (
                        WHERE is_active = TRUE
                          AND reorder_level > 0
                          AND quantity_on_hand <= reorder_level
                    ) AS low_stock_count
                FROM products;
                """
            )
        ).first()

        sales_summary = conn.execute(
            text(
                """
                SELECT
                    COUNT(*) AS today_sale_count,
                    COALESCE(SUM(total_cents), 0) AS today_revenue_cents
                FROM sales
                WHERE created_at::date = CURRENT_DATE;
                """
            )
        ).first()

        online_order_summary = conn.execute(
            text(
                """
                SELECT
                    COUNT(*) FILTER (
                        WHERE status IN ('pending_packaging', 'packaged')
                    ) AS pending_fulfillment_count
                FROM online_orders;
                """
            )
        ).first()

        low_stock_items = conn.execute(
            text(
                """
                SELECT id, sku, name, quantity_on_hand, reorder_level
                FROM products
                WHERE is_active = TRUE
                  AND reorder_level > 0
                  AND quantity_on_hand <= reorder_level
                ORDER BY quantity_on_hand ASC, name ASC
                LIMIT 8;
                """
            )
        )

    return {
        "products": dict(product_summary._mapping),
        "sales": dict(sales_summary._mapping),
        "online_orders": dict(online_order_summary._mapping),
        "low_stock_items": [dict(row._mapping) for row in low_stock_items],
    }


@app.get("/inventory/history/{product_id}")
def get_inventory_history(product_id: int):
    with engine.connect() as conn:
        result = conn.execute(
            text(
                """
                SELECT *
                FROM inventory_transactions
                WHERE product_id = :product_id
                ORDER BY created_at DESC;
                """
            ),
            {"product_id": product_id},
        )

        return [
            dict(row._mapping)
            for row in result
        ]
    

@app.get("/sales")
def get_sales():
    with engine.connect() as conn:
        result = conn.execute(
            text(
                """
                SELECT *
                FROM sales
                ORDER BY created_at DESC;
                """
            )
        )

        return [dict(row._mapping) for row in result]
    

@app.get("/sales/today")
def get_today_sales():
    with engine.connect() as conn:
        sales = conn.execute(
            text(
                """
                SELECT *
                FROM sales
                WHERE created_at::date = CURRENT_DATE
                ORDER BY created_at DESC;
                """
            )
        )

        summary = conn.execute(
            text(
                """
                SELECT
                    COUNT(*) AS sale_count,
                    COALESCE(SUM(total_cents), 0) AS total_cents
                FROM sales
                WHERE created_at::date = CURRENT_DATE;
                """
            )
        ).first()

        return {
            "summary": dict(summary._mapping),
            "sales": [dict(row._mapping) for row in sales],
        }

@app.get("/sales/range")
def get_sales_range(start_date: str, end_date: str):
    with engine.connect() as conn:
        sales = conn.execute(
            text(
                """
                SELECT
                    sales.*,
                    COALESCE(
                        sales.customer_name,
                        online_orders.customer_name,
                        online_orders.shipping_name
                    ) AS display_customer_name,
                    online_orders.id AS online_order_id,
                    online_orders.status AS online_order_status
                FROM sales
                LEFT JOIN online_orders ON online_orders.sale_id = sales.id
                WHERE sales.created_at >= :start_date
                  AND sales.created_at < (CAST(:end_date AS date) + INTERVAL '1 day')
                ORDER BY sales.created_at DESC;
                """
            ),
            {
                "start_date": start_date,
                "end_date": end_date,
            },
        )

        summary = conn.execute(
            text(
                """
                SELECT
                    COUNT(*) AS sale_count,
                    COALESCE(SUM(total_cents), 0) AS total_cents
                FROM sales
                WHERE created_at >= :start_date
                  AND created_at < (CAST(:end_date AS date) + INTERVAL '1 day');
                """
            ),
            {
                "start_date": start_date,
                "end_date": end_date,
            },
        ).first()

        return {
            "summary": dict(summary._mapping),
            "sales": [dict(row._mapping) for row in sales],
        }
    
    
@app.get("/sales/search")
def search_sales(query: str, field: str = "all"):
    search = f"%{query.strip()}%"

    allowed_fields = {
        "all",
        "order_number",
        "customer",
        "tracking",
        "type",
        "product",
    }

    if field not in allowed_fields:
        raise HTTPException(status_code=400, detail="Invalid search field")

    if not query.strip():
        return {
            "summary": {"sale_count": 0, "total_cents": 0},
            "sales": [],
        }

    conditions = {
        "order_number": """
            (
                sales.order_number ILIKE :search
                OR sales.id::text ILIKE :search
                OR online_orders.id::text ILIKE :search
            )
        """,
        "customer": """
            (
                sales.customer_name ILIKE :search
                OR online_orders.customer_name ILIKE :search
                OR online_orders.customer_email ILIKE :search
                OR online_orders.shipping_name ILIKE :search
            )
        """,
        "tracking": """
            (
                online_orders.tracking_id ILIKE :search
                OR online_orders.carrier ILIKE :search
            )
        """,
        "type": """
            (
                (:query_lower = 'online' AND online_orders.id IS NOT NULL)
                OR (:query_lower = 'pos' AND online_orders.id IS NULL)
            )
        """,
        "product": """
            EXISTS (
                SELECT 1
                FROM sale_items
                JOIN products ON products.id = sale_items.product_id
                WHERE sale_items.sale_id = sales.id
                  AND products.name ILIKE :search
            )
        """,
    }

    if field == "all":
        where_clause = " OR ".join(f"({condition})" for condition in conditions.values())
    else:
        where_clause = conditions[field]

    with engine.connect() as conn:
        sales = conn.execute(
            text(
                f"""
                SELECT
                    sales.*,
                    COALESCE(
                        sales.customer_name,
                        online_orders.customer_name,
                        online_orders.shipping_name
                    ) AS display_customer_name,
                    online_orders.id AS online_order_id,
                    online_orders.status AS online_order_status
                FROM sales
                LEFT JOIN online_orders ON online_orders.sale_id = sales.id
                WHERE {where_clause}
                ORDER BY sales.created_at DESC
                LIMIT 100;
                """
            ),
            {
                "search": search,
                "query_lower": query.strip().lower(),
            },
        )

        sale_rows = [dict(row._mapping) for row in sales]
        total_cents = sum(row["total_cents"] for row in sale_rows)

        return {
            "summary": {
                "sale_count": len(sale_rows),
                "total_cents": total_cents,
            },
            "sales": sale_rows,
        }


@app.get("/sales/{sale_id}")
def get_sale(sale_id: int):
    with engine.connect() as conn:
        sale = conn.execute(
            text("SELECT * FROM sales WHERE id = :id;"),
            {"id": sale_id},
        ).first()

        if sale is None:
            raise HTTPException(status_code=404, detail="Sale not found")

        items = conn.execute(
            text(
                """
                SELECT
                    sale_items.id,
                    sale_items.sale_id,
                    sale_items.product_id,
                    products.name,
                    sale_items.quantity,
                    sale_items.price_cents
                FROM sale_items
                JOIN products ON products.id = sale_items.product_id
                WHERE sale_items.sale_id = :sale_id
                ORDER BY sale_items.id;
                """
            ),
            {"sale_id": sale_id},
        )

        online_order = conn.execute(
            text(
                """
                SELECT *
                FROM online_orders
                WHERE sale_id = :sale_id;
                """
            ),
            {"sale_id": sale_id},
        ).first()

        return {
            "sale": dict(sale._mapping),
            "items": [dict(row._mapping) for row in items],
            "online_order": dict(online_order._mapping) if online_order else None,
        }
    

@app.get("/inventory/low-stock")
def get_low_stock_products():
    with engine.connect() as conn:
        result = conn.execute(
            text(
                """
                SELECT *
                FROM products
                WHERE is_active = TRUE
                  AND reorder_level > 0
                  AND quantity_on_hand <= reorder_level
                ORDER BY quantity_on_hand ASC, name ASC;
                """
            )
        )

        return [dict(row._mapping) for row in result]
    

@app.post("/inventory/adjust")
def adjust_inventory(adjustment: InventoryAdjustment):
    with engine.begin() as conn:

        product = conn.execute(
            text(
                """
                SELECT *
                FROM products
                WHERE id = :id;
                """
            ),
            {"id": adjustment.product_id},
        ).first()

        if product is None:
            raise HTTPException(
                status_code=404,
                detail="Product not found"
            )

        conn.execute(
            text(
                """
                INSERT INTO inventory_transactions
                (
                    product_id,
                    quantity_change,
                    reason,
                    notes
                )
                VALUES
                (
                    :product_id,
                    :quantity_change,
                    :reason,
                    :notes
                );
                """
            ),
            adjustment.model_dump(),
        )

        conn.execute(
            text(
                """
                UPDATE products
                SET quantity_on_hand =
                    quantity_on_hand + :quantity_change,
                    updated_at = CURRENT_TIMESTAMP
                WHERE id = :product_id;
                """
            ),
            adjustment.model_dump(),
        )

        updated_product = conn.execute(
            text(
                """
                SELECT *
                FROM products
                WHERE id = :id;
                """
            ),
            {"id": adjustment.product_id},
        ).first()

    return dict(updated_product._mapping)


@app.post("/products")
def create_product(product: ProductCreate):
    with engine.begin() as conn:
        result = conn.execute(
            text(
                """
                INSERT INTO products
                (
                    sku,
                    barcode,
                    name,
                    category,
                    description,
                    price_cents,
                    cost_cents,
                    quantity_on_hand,
                    reorder_level,
                    image_url,
                    public_description,
                    is_public
                )
                VALUES
                (
                    :sku,
                    :barcode,
                    :name,
                    :category,
                    :description,
                    :price_cents,
                    :cost_cents,
                    :quantity_on_hand,
                    :reorder_level,
                    :image_url,
                    :public_description,
                    :is_public
                )
                RETURNING *;
                """
            ),
            product.model_dump(),
        ).first()

    return dict(result._mapping)


@app.put("/products/{product_id}")
def update_product(product_id: int, product: ProductUpdate):
    fields = product.model_dump(exclude_unset=True)

    if not fields:
        raise HTTPException(
            status_code=400,
            detail="No update fields provided"
        )

    set_clause = ", ".join(
        [f"{key} = :{key}" for key in fields.keys()]
    )

    query = f"""
        UPDATE products
        SET {set_clause},
            updated_at = CURRENT_TIMESTAMP
        WHERE id = :id
        RETURNING *;
    """

    fields["id"] = product_id

    with engine.begin() as conn:
        result = conn.execute(
            text(query),
            fields
        ).first()

    if result is None:
        raise HTTPException(
            status_code=404,
            detail="Product not found"
        )

    return dict(result._mapping)


@app.post("/products/{product_id}/image")
async def upload_product_image(product_id: int, file: UploadFile = File(...)):
    with engine.connect() as conn:
        product = conn.execute(
            text("SELECT * FROM products WHERE id = :id;"),
            {"id": product_id},
        ).first()

    if product is None:
        raise HTTPException(status_code=404, detail="Product not found")

    _, extension = os.path.splitext(file.filename or "")
    extension = extension.lower()

    if extension not in ALLOWED_IMAGE_EXTENSIONS:
        raise HTTPException(
            status_code=400,
            detail="Image must be a JPG, PNG, WebP, or GIF file",
        )

    if not file.content_type or not file.content_type.startswith("image/"):
        raise HTTPException(status_code=400, detail="Uploaded file must be an image")

    contents = await file.read()

    if len(contents) > MAX_IMAGE_BYTES:
        raise HTTPException(status_code=400, detail="Image must be 5 MB or smaller")

    filename = f"product-{product_id}-{uuid.uuid4().hex}{extension}"
    disk_path = os.path.join(PRODUCT_UPLOAD_DIR, filename)
    image_url = f"/uploads/products/{filename}"

    with open(disk_path, "wb") as image_file:
        image_file.write(contents)

    with engine.begin() as conn:
        updated_product = conn.execute(
            text(
                """
                UPDATE products
                SET image_url = :image_url,
                    updated_at = CURRENT_TIMESTAMP
                WHERE id = :id
                RETURNING *;
                """
            ),
            {"id": product_id, "image_url": image_url},
        ).first()

    return dict(updated_product._mapping)


@app.get("/products/{product_id}/images")
def get_product_images(product_id: int):
    with engine.connect() as conn:
        product = conn.execute(
            text("SELECT * FROM products WHERE id = :id;"),
            {"id": product_id},
        ).first()

        if product is None:
            raise HTTPException(status_code=404, detail="Product not found")

        result = conn.execute(
            text(
                """
                SELECT *
                FROM product_images
                WHERE product_id = :product_id
                ORDER BY created_at ASC, id ASC;
                """
            ),
            {"product_id": product_id},
        )

        return [dict(row._mapping) for row in result]


@app.post("/products/{product_id}/images")
async def upload_product_gallery_image(product_id: int, file: UploadFile = File(...)):
    with engine.connect() as conn:
        product = conn.execute(
            text("SELECT * FROM products WHERE id = :id;"),
            {"id": product_id},
        ).first()

    if product is None:
        raise HTTPException(status_code=404, detail="Product not found")

    _, extension = os.path.splitext(file.filename or "")
    extension = extension.lower()

    if extension not in ALLOWED_IMAGE_EXTENSIONS:
        raise HTTPException(
            status_code=400,
            detail="Image must be a JPG, PNG, WebP, or GIF file",
        )

    if not file.content_type or not file.content_type.startswith("image/"):
        raise HTTPException(status_code=400, detail="Uploaded file must be an image")

    contents = await file.read()

    if len(contents) > MAX_IMAGE_BYTES:
        raise HTTPException(status_code=400, detail="Image must be 5 MB or smaller")

    filename = f"product-{product_id}-gallery-{uuid.uuid4().hex}{extension}"
    disk_path = os.path.join(PRODUCT_UPLOAD_DIR, filename)
    image_url = f"/uploads/products/{filename}"

    with open(disk_path, "wb") as image_file:
        image_file.write(contents)

    with engine.begin() as conn:
        result = conn.execute(
            text(
                """
                INSERT INTO product_images (product_id, image_url)
                VALUES (:product_id, :image_url)
                RETURNING *;
                """
            ),
            {"product_id": product_id, "image_url": image_url},
        ).first()

    return dict(result._mapping)


@app.delete("/product-images/{image_id}")
def delete_product_gallery_image(image_id: int):
    with engine.begin() as conn:
        result = conn.execute(
            text(
                """
                DELETE FROM product_images
                WHERE id = :id
                RETURNING *;
                """
            ),
            {"id": image_id},
        ).first()

    if result is None:
        raise HTTPException(status_code=404, detail="Product image not found")

    return {"deleted": True, "image": dict(result._mapping)}


@app.post("/sales")
def create_sale(sale: SaleCreate):
    if not sale.items:
        raise HTTPException(status_code=400, detail="Sale must contain at least one item")

    with engine.begin() as conn:
        settings = get_app_settings(conn)
        subtotal_cents = 0
        sale_items_data = []

        for item in sale.items:
            product = conn.execute(
                text("SELECT * FROM products WHERE id = :id;"),
                {"id": item.product_id},
            ).first()

            if product is None:
                raise HTTPException(
                    status_code=404,
                    detail=f"Product {item.product_id} not found"
                )

            product_data = dict(product._mapping)

            if item.quantity <= 0:
                raise HTTPException(
                    status_code=400,
                    detail="Quantity must be greater than zero"
                )

            if product_data["quantity_on_hand"] < item.quantity:
                raise HTTPException(
                    status_code=400,
                    detail=f"Not enough inventory for {product_data['name']}"
                )

            line_total = product_data["price_cents"] * item.quantity
            subtotal_cents += line_total

            sale_items_data.append({
                "product_id": item.product_id,
                "quantity": item.quantity,
                "price_cents": product_data["price_cents"],
                "name": product_data["name"],
            })

        tax_cents = (
            calculate_tax_cents(subtotal_cents, settings["tax_rate_percent"])
            if settings["tax_enabled"]
            else 0
        )
        tax_rate_percent = settings["tax_rate_percent"] if settings["tax_enabled"] else 0
        unrounded_total_cents = subtotal_cents + tax_cents
        rounding_adjustment_cents = calculate_rounding_adjustment_cents(
            unrounded_total_cents,
            settings["pos_rounding_mode"],
        )
        total_cents = unrounded_total_cents + rounding_adjustment_cents

        sale_row = conn.execute(
            text(
                """
                INSERT INTO sales (
                    customer_name,
                    payment_type,
                    subtotal_cents,
                    tax_cents,
                    tax_rate_percent,
                    rounding_adjustment_cents,
                    total_cents,
                    order_number
                )
                VALUES (
                    :customer_name,
                    :payment_type,
                    :subtotal_cents,
                    :tax_cents,
                    :tax_rate_percent,
                    :rounding_adjustment_cents,
                    :total_cents,
                    :order_number
                )
                RETURNING *;
                """
            ),
            {
                "customer_name": sale.customer_name.strip() if sale.customer_name else None,
                "payment_type": sale.payment_type or "cash",
                "subtotal_cents": subtotal_cents,
                "tax_cents": tax_cents,
                "tax_rate_percent": tax_rate_percent,
                "rounding_adjustment_cents": rounding_adjustment_cents,
                "total_cents": total_cents,
                "order_number": generate_order_number(conn),
            },
        ).first()

        sale_id = sale_row._mapping["id"]

        for item in sale_items_data:
            conn.execute(
                text(
                    """
                    INSERT INTO sale_items
                    (sale_id, product_id, quantity, price_cents)
                    VALUES
                    (:sale_id, :product_id, :quantity, :price_cents);
                    """
                ),
                {
                    "sale_id": sale_id,
                    "product_id": item["product_id"],
                    "quantity": item["quantity"],
                    "price_cents": item["price_cents"],
                },
            )

            conn.execute(
                text(
                    """
                    UPDATE products
                    SET quantity_on_hand = quantity_on_hand - :quantity,
                        updated_at = CURRENT_TIMESTAMP
                    WHERE id = :product_id;
                    """
                ),
                {
                    "quantity": item["quantity"],
                    "product_id": item["product_id"],
                },
            )

            conn.execute(
                text(
                    """
                    INSERT INTO inventory_transactions
                    (product_id, quantity_change, reason, notes)
                    VALUES
                    (:product_id, :quantity_change, 'Sale', :notes);
                    """
                ),
                {
                    "product_id": item["product_id"],
                    "quantity_change": -item["quantity"],
                    "notes": f"Sale #{sale_id}",
                },
            )

    return {
        "id": sale_id,
        "order_number": sale_row._mapping["order_number"],
        "customer_name": sale_row._mapping["customer_name"],
        "payment_type": sale_row._mapping["payment_type"],
        "subtotal_cents": sale_row._mapping["subtotal_cents"],
        "tax_cents": sale_row._mapping["tax_cents"],
        "tax_rate_percent": str(sale_row._mapping["tax_rate_percent"]),
        "rounding_adjustment_cents": sale_row._mapping["rounding_adjustment_cents"],
        "total_cents": sale_row._mapping["total_cents"],
        "items": sale_items_data,
    }



@app.delete("/products/{product_id}")
def delete_product(product_id: int):
    with engine.begin() as conn:
        result = conn.execute(
            text(
                """
                UPDATE products
                SET is_active = FALSE,
                    updated_at = CURRENT_TIMESTAMP
                WHERE id = :id
                RETURNING *;
                """
            ),
            {"id": product_id},
        ).first()

    if result is None:
        raise HTTPException(status_code=404, detail="Product not found")

    return {"deleted": True, "product": dict(result._mapping)}
